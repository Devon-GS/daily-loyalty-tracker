import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, Font
import os


# Files
monthly_Shifts = 'Files/Monthly Shift Summary.csv'
tender_file = 'Files/Tender Provider Detail - Transaction Date Range.csv'

# List of filteres data
shifts = []
tender = []
batches = {}
shift_total = {}

# MONTHLY SHIFT NUMBERS
monthly_shifts = pd.read_csv(monthly_Shifts, skip_blank_lines=True).fillna(value='None')

filla_shifts = monthly_shifts.iloc[:, [1, 3]].values.tolist()

for x in filla_shifts:
    if x[0] != 'None':
        shifts.append([str(int(x[0])), str(int(x[1]))])
        
for x in shifts:
    if x[1] not in batches:
        batches[x[1]] = [x[0]]
    else:
        batches[x[1]].append(x[0])

        
# MONTHLY TENDERS
monthly_tender = pd.read_csv(tender_file, skip_blank_lines=True).fillna(value='None')

filla_trans = monthly_tender.iloc[:, [0, 5, 8]].values.tolist()

for x in filla_trans:
    if x[0] != 'None':
        tender_desc = x[0].strip()
        shift_number = str(int(x[1]))
        amount = float(x[2])
        
        if tender_desc == 'Loyalty':
            tender.append([shift_number, amount])
            
for k, v in batches.items():
    for x in tender:
        if x[0] in v:
            if k not in shift_total:
                shift_total[k] = float(x[1])	
            else:
                shift_total[k] += float(x[1])
                
# CREATE Workbook

# Styling setup
heading_format = NamedStyle(name="heading_format")
heading_format.font = Font(bold=True, u='single')
heading_format.alignment = Alignment(horizontal='center')

wb = Workbook()
ws = wb.active

ws['A1'] = 'Batch Number'
ws['B1'] = 'Total'

# Loop through dict and copy info to workbook
i = 2

for x, y in shift_total.items():
    ws[f'A{i}'] = x
    ws[f'B{i}'] = y
    
    i += 1

 # Styling
ws['A1'].style = heading_format
ws['B1'].style = heading_format

ws.column_dimensions['A'].width = 12.64
ws.column_dimensions['B'].width = 7.97

# Save the file
wb.save("Daily Loyalty.xlsx")
wb.close()

os.system('start "EXCEL.EXE" "Daily Loyalty.xlsx"')